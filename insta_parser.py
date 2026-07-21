#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Instagram lead generation tool: finds English-speaking coaches, consultants and
other service experts and extracts every contact channel from their profile.

Architecture: the data source (Apify / HikerAPI) is decoupled from the pipeline.
Both sources are normalized into a single Profile object, then everything else
is shared:
    source -> normalize -> filters -> contact extraction -> dedup -> Excel

Usage:
    pip install -r requirements.txt
    export APIFY_TOKEN="..."      # for a small test run
    export HIKER_TOKEN="..."      # for volume (HikerAPI, cheaper per request)
    python insta_parser.py --source apify --limit 50

A note on schemas: exact field names returned by Apify actors vary between
actor versions. normalize_profile() reads every known variant of each field,
but if an actor is updated, check its Output tab in Apify Store and add the
new key to the mapping if needed.
"""

import os
import re
import sys
import time
import argparse
import logging
from dataclasses import dataclass, field
from typing import Optional, Iterable
from urllib.parse import unquote, urlsplit, urlunsplit

import requests
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("collector")


# Keywords used both for Instagram search and for the is_expert() bio/name/category check.
EXPERT_KEYWORDS = [
    "coach", "life coach", "business coach", "consultant", "mentor", "therapist",
    "nutritionist", "dietitian", "personal trainer", "marketing consultant",
    "seo expert", "copywriter", "web designer", "ux designer", "brand strategist",
    "career coach", "sales coach", "financial advisor", "business strategist",
    "social media manager", "content creator", "astrologer", "stylist",
    "mindset coach", "productivity coach",
]

# Signals of relevance from the brief. These never reject a profile, they only
# add to relevance_score (see score_relevance()).
BONUS_KEYWORDS = [
    "book a call", "free consultation", "dm me", "work with me", "coaching",
    "mentorship", "1:1", "session", "enroll", "course", "masterclass",
    "webinar", "free guide", "checklist", "program", "apply now", "limited spots",
]

# Anything that signals a business/shop rather than an individual expert.
EXCLUDE_KEYWORDS = [
    "shop", "store", "boutique", "delivery", "restaurant", "cafe", "salon",
    "brand", "official", "agency", "real estate", "e-commerce", "dropshipping",
    "marketplace", "wholesale", "retail", "factory", "manufacturer",
]

MIN_FOLLOWERS = 1000
MAX_FOLLOWERS = 30000
MAX_FOLLOWING = 2500

# Handles that are never a real Telegram username (common false positives from @mentions,
# plus noise picked up once we started scanning arbitrary web pages: share widgets,
# language/proxy bot deep links, instant-view links, etc.)
HANDLE_STOPLIST = {
    "gmail", "mail", "instagram", "insta", "youtube", "youtu", "tiktok",
    "whatsapp", "wa", "vk", "facebook", "fb", "twitter", "x", "threads",
    "share", "s", "widget", "addstickers", "iv", "proxy", "setlanguage",
}

# Column order and headers of the exported Excel file.
EXCEL_COLUMNS = [
    ("username", "instagram_username"),
    ("instagram_url", "instagram_url"),
    ("full_name", "full_name"),
    ("biography", "biography"),
    ("followers", "followers"),
    ("following", "following"),
    ("category", "category"),
    ("matched_keyword", "matched_keyword"),
    ("relevance_score", "relevance_score"),
    ("email", "email"),
    ("whatsapp", "whatsapp"),
    ("telegram_username", "telegram_username"),
    ("telegram_url", "telegram_url"),
    ("booking_link", "booking_link"),
    ("website", "website"),
    ("notes", "notes"),
]


def matches_exclusion_markers(profile: "Profile", markers: list[str]) -> bool:
    """Case-insensitive substring match of `markers` against bio/name/category.

    Optional geo/keyword exclusion filter, off by default (empty markers list).
    Pass e.g. --exclude-countries kyiv ukraine to drop profiles mentioning them.
    """
    if not markers:
        return False
    text = profile.search_text
    return any(marker.lower() in text for marker in markers)


@dataclass
class Profile:
    username: str
    full_name: str = ""
    biography: str = ""
    followers: int = 0
    following: int = 0
    category: str = ""
    external_url: str = ""                              # all bio links joined into one blob
    external_urls: list[str] = field(default_factory=list)  # same links, kept individually
    provider_email: str = ""                             # business/public email from the API, if any
    matched_keyword: str = ""
    email: str = ""
    whatsapp: str = ""
    telegram_username: str = ""
    booking_link: str = ""
    website: str = ""
    notes: str = ""
    relevance_score: int = 0
    raw: dict = field(default_factory=dict, repr=False)

    @property
    def instagram_url(self) -> str:
        return f"https://www.instagram.com/{self.username}/"

    @property
    def telegram_url(self) -> str:
        return f"https://t.me/{self.telegram_username}" if self.telegram_username else ""

    @property
    def search_text(self) -> str:
        """All free text fields joined for keyword matching."""
        return " ".join([self.full_name, self.biography, self.category]).lower()


def _to_int(v) -> int:
    if isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, dict):  # shape like {"count": 1234}
        return _to_int(v.get("count"))
    if isinstance(v, str):
        digits = re.sub(r"[^\d]", "", v)
        return int(digits) if digits else 0
    return 0


def _first(d: dict, *keys):
    """Return the first non-empty value among the given keys."""
    for k in keys:
        if k in d and d[k] not in (None, "", []):
            return d[k]
    return None


def normalize_profile(d: dict) -> Optional[Profile]:
    """Convert a raw Apify or HikerAPI item into the common Profile model."""
    if not isinstance(d, dict):
        return None

    username = _first(d, "username", "ownerUsername", "user_name")
    if not username:
        # HikerAPI sometimes nests the profile under "user".
        if isinstance(d.get("user"), dict):
            return normalize_profile(d["user"])
        return None

    # Collect every bio link, not just the single external_url field: Apify actors
    # may return a list (channel, booking page, personal site, ...).
    external_parts: list[str] = []
    single = _first(d, "external_url", "externalUrl", "bio_link_url", "website")
    if single:
        external_parts.append(str(single))
    for key in ("external_urls", "externalUrls", "bio_links", "biolinks"):
        ext_list = d.get(key)
        if isinstance(ext_list, list):
            for item in ext_list:
                url = item.get("url") if isinstance(item, dict) else item
                if url:
                    external_parts.append(str(url))

    # Field name unconfirmed against a live response, verify against the actor's
    # current Output tab and extend this list if it uses a different key.
    provider_email = _first(d, "businessEmail", "business_email", "publicEmail", "public_email")

    return Profile(
        username=str(username).lstrip("@").strip(),
        full_name=str(_first(d, "full_name", "fullName", "name") or "").strip(),
        biography=str(_first(d, "biography", "bio", "biography_with_entities") or "").strip(),
        followers=_to_int(_first(d, "follower_count", "followersCount", "followers",
                                 "edge_followed_by")),
        following=_to_int(_first(d, "following_count", "followsCount", "follows_count",
                                 "followingCount", "edge_follow")),
        category=str(_first(d, "category", "category_name", "businessCategoryName",
                            "business_category_name") or "").strip(),
        external_url=" ".join(external_parts).strip(),
        external_urls=external_parts,
        provider_email=str(provider_email or "").strip(),
        raw=d,
    )


# t.me/username, t.me/+invite, telegram.me/username, tg://resolve?domain=username
TME_RE = re.compile(
    r"(?:https?://)?(?:www\.)?(?:t\.me|telegram\.me|telegram\.dog)/"
    r"(\+?[A-Za-z0-9_./-]{3,})",
    re.IGNORECASE,
)
TG_DEEPLINK_RE = re.compile(r"tg://resolve\?domain=([A-Za-z0-9_]{4,32})", re.IGNORECASE)

# An @handle next to the word "telegram/tg" (in English or Russian) is a reliable signal.
CONTEXT_TG_RE = re.compile(
    r"(?:telegram|телеграм|телега|тelega|тг|tg)\s*[:\-—>|]*\s*@?([A-Za-z0-9_]{5,32})",
    re.IGNORECASE,
)
# Any @handle, used only if the bio mentions Telegram at all.
BARE_HANDLE_RE = re.compile(r"@([A-Za-z0-9_]{5,32})")

TG_MENTION_RE = re.compile(r"telegram|телеграм|телега|\bтг\b|\btg\b", re.IGNORECASE)

TME_REQUEST_TIMEOUT = 15
TME_REQUEST_DELAY = 0.3  # be polite to t.me between requests

TME_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
}

# Markup signals confirmed on real t.me pages (t.me/telegram -> "subscribers",
# t.me/pythonchatru -> "members", action button "View in Telegram"/"Preview channel").
# Kept bilingual since t.me page copy can render in the channel owner's locale.
CHANNEL_EXTRA_MARKERS = ["subscribers", "подписчик", "members", "участник"]
CHANNEL_BUTTON_LABELS = {
    "view in telegram", "preview channel", "join channel", "join group",
    "открыть в telegram",
}
# "Send Message" with no subscriber/member count is the personal-account signal.
# t.me shows this same button for usernames that don't exist at all, so
# classify_tme_target() only trusts it once tgme_page_photo confirms the
# account is real. See that function for why.
PERSONAL_BUTTON_LABELS = {"send message", "написать"}

# Domains treated as link-in-bio aggregators; their pages get scraped for contacts
# (email/WhatsApp/Telegram/booking) but are never reported as the profile's own "website".
LINK_AGGREGATOR_DOMAINS = (
    "linktr.ee", "taplink", "linktree", "lnk.bio", "beacons", "hipolink", "mssg.me",
    "bio.link", "linkin.bio", "milkshake", "campsite.bio", "koji.to", "stan.store",
)

TME_ACTION_BUTTON_RE = re.compile(r'tgme_action_button_new[^"]*"[^>]*>([^<]*)</a>')
TME_PAGE_EXTRA_RE = re.compile(r'tgme_page_extra">([^<]*)')


def _clean_handle(h: str) -> Optional[str]:
    h = h.strip().strip("/").lstrip("@")
    if not h:
        return None
    if h.lower() in HANDLE_STOPLIST:
        return None
    return h


# Cache of t.me page classifications: handle(lower) -> "bot"/"channel"/"personal"/"unknown"
_tme_type_cache: dict[str, str] = {}


def classify_tme_target(handle: str) -> str:
    """Classify t.me/<handle> as "bot" / "channel" / "personal" / "unknown".

    A name ending in "bot" is always a bot, no request needed. Otherwise
    fetch https://t.me/<handle> and read the markup.

    t.me renders the same "Send Message" button for a username that doesn't
    exist at all as it does for a real personal account, so the button alone
    is not proof the handle is real. The one marker that reliably differs is
    "tgme_page_photo" (an avatar block), rendered only when an account,
    channel or bot actually exists behind the handle. If it's missing, the
    result is "unknown" regardless of anything else on the page.

    Once tgme_page_photo confirms the handle is real:
      - "Start Bot" button or "monthly users" in tgme_page_extra -> bot
      - subscriber/member count, or a "View in Telegram"/"Preview channel"
        button -> channel or group
      - "Send Message" button -> personal account

    Results are cached so the same handle is never fetched twice.
    """
    key = handle.lower()
    if key.endswith("bot"):
        return "bot"
    if key in _tme_type_cache:
        return _tme_type_cache[key]

    kind = "unknown"
    try:
        time.sleep(TME_REQUEST_DELAY)
        r = requests.get(f"https://t.me/{handle}", timeout=TME_REQUEST_TIMEOUT,
                         headers=TME_HEADERS)
        html = r.text
        if "tgme_page_wrap" in html and "tgme_page_photo" in html:
            btn_m = TME_ACTION_BUTTON_RE.search(html)
            btn = btn_m.group(1).strip().lower() if btn_m else ""
            extra = " ".join(t.strip() for t in TME_PAGE_EXTRA_RE.findall(html)).lower()

            if "start bot" in btn or "monthly users" in extra:
                kind = "bot"
            elif any(m in extra for m in CHANNEL_EXTRA_MARKERS) or btn in CHANNEL_BUTTON_LABELS:
                kind = "channel"
            elif btn in PERSONAL_BUTTON_LABELS:
                kind = "personal"
    except Exception as e:
        log.debug("t.me check %s: %s", handle, e)

    _tme_type_cache[key] = kind
    return kind


def collect_tg_candidates(text_sources: Iterable[tuple[str, str]], own: str) -> list[tuple[str, str]]:
    """Collect every t.me handle / @-mention found in the given texts (bio, links, ...).

    Returns (handle_without_at, source_label) pairs in priority order, deduped.
    Invite links (t.me/+xxx) are skipped: they never resolve to a personal handle.
    """
    candidates: list[tuple[str, str]] = []
    seen: set[str] = set()

    def add(raw_handle: Optional[str], label: str):
        h = _clean_handle(raw_handle) if raw_handle else None
        if h and h.lower() != own and h.lower() not in seen:
            seen.add(h.lower())
            candidates.append((h, label))

    for src, txt in text_sources:
        if not txt:
            continue
        for m in TME_RE.finditer(txt):
            raw = m.group(1).split("/")[0]
            if raw.startswith("+") or raw.lower().startswith("joinchat"):
                continue  # invite link to a channel/group, not a personal handle
            add(raw, f"t.me ({src})")
        for m in TG_DEEPLINK_RE.finditer(txt):
            add(m.group(1), f"tg://resolve ({src})")

        m = CONTEXT_TG_RE.search(txt)
        if m:
            add(m.group(1), f"context (tg/telegram) [{src}]")

        # "Any @word counts once telegram is mentioned anywhere" is only safe on a
        # short, tightly-scoped text like an Instagram bio. On a fetched HTML page
        # a stray "telegram" mention (e.g. a share-icon widget) sits alongside
        # completely unrelated @-tokens (CSS "@import", JSON-LD "@graph", ...) that
        # can coincidentally be someone else's real t.me handle, so this fallback
        # is restricted to the bio source only.
        if src == "bio" and TG_MENTION_RE.search(txt):
            for m in BARE_HANDLE_RE.finditer(txt):
                add(m.group(1), f"@handle (telegram mentioned) [{src}]")

    return candidates


def extract_telegram(profile: Profile) -> tuple[Optional[str], str, dict]:
    """Return (personal handle without @, note, candidate stats) from bio + bio link.

    Takes the first candidate whose t.me page classifies as personal; bots and
    channels are discarded.
    """
    stats = {"candidates": 0, "bot": 0, "channel": 0, "unknown": 0, "personal": 0}
    own = profile.username.lower()
    sources = [("bio-link", profile.external_url or ""), ("bio", profile.biography or "")]

    for handle, _label in collect_tg_candidates(sources, own):
        stats["candidates"] += 1
        kind = classify_tme_target(handle)
        stats[kind] = stats.get(kind, 0) + 1
        if kind == "personal":
            return handle, "found in bio", stats

    return None, "", stats


# Cache of fetched pages, so the same link is never fetched twice across a whole run.
_page_cache: dict[str, Optional[str]] = {}

MAX_RESPONSE_BYTES = 2_000_000       # don't read past this from any single page
MAX_EXTRA_WEBSITE_PAGES = 2          # /contact, /contact-us, /about, capped per profile
CONTACT_PAGE_PATHS = ("/contact", "/contact-us", "/about")


def fetch_page(url: str) -> Optional[str]:
    """GET a URL with a browser UA, capped response size, cached by URL so the same
    link is never fetched twice in a run. Errors/timeouts are swallowed silently:
    callers just get None back, meaning "no contact found here"."""
    if url in _page_cache:
        return _page_cache[url]

    html: Optional[str] = None
    try:
        time.sleep(TME_REQUEST_DELAY)
        r = requests.get(url, timeout=TME_REQUEST_TIMEOUT, headers=TME_HEADERS, stream=True)
        chunks = []
        size = 0
        for chunk in r.iter_content(8192):
            chunks.append(chunk)
            size += len(chunk)
            if size >= MAX_RESPONSE_BYTES:
                break
        html = b"".join(chunks).decode(r.encoding or "utf-8", errors="ignore")
    except Exception as e:
        log.debug("fetch %s: %s", url, e)

    _page_cache[url] = html
    return html


def _empty_scan_result() -> dict:
    return {
        "email": "", "whatsapp": "", "telegram_username": "", "booking_link": "",
        "tg_stats": {"candidates": 0, "bot": 0, "channel": 0, "unknown": 0, "personal": 0},
        "fetched": False,
    }


def scan_page_for_contacts(html: str, own_username: str, label: str) -> dict:
    """Scan a fetched page's HTML for every contact channel: email (incl. mailto:),
    WhatsApp (wa.me / api.whatsapp.com), a personal Telegram handle, and a booking link."""
    result = _empty_scan_result()
    result["fetched"] = True

    mailto_m = MAILTO_RE.search(html)
    if mailto_m:
        candidate = unquote(mailto_m.group(1)).split("?")[0]
        m = EMAIL_RE.search(candidate)
        if m:
            result["email"] = m.group(0)
    if not result["email"]:
        m = EMAIL_RE.search(html)
        if m:
            result["email"] = m.group(0)

    result["whatsapp"] = extract_whatsapp(html)

    for m in URL_IN_HTML_RE.finditer(html):
        if classify_link(m.group(0)) == "booking":
            result["booking_link"] = m.group(0)
            break

    for handle, _src in collect_tg_candidates([(label, html)], own_username):
        result["tg_stats"]["candidates"] += 1
        kind = classify_tme_target(handle)
        result["tg_stats"][kind] = result["tg_stats"].get(kind, 0) + 1
        if kind == "personal":
            result["telegram_username"] = handle
            break

    return result


def resolve_link_aggregator(url: str, own_username: str = "") -> dict:
    """Fetch a link-in-bio aggregator page (Linktree, Taplink, bio.link, ...) and scan
    it for every contact channel, not just Telegram, since experts often put their
    email/WhatsApp/Calendly link there instead of directly in the Instagram bio."""
    if not url:
        return _empty_scan_result()
    html = fetch_page(url)
    if html is None:
        return _empty_scan_result()
    return scan_page_for_contacts(html, own_username, f"linktree ({url})")


def resolve_website(url: str, own_username: str = "") -> dict:
    """Fetch a plain website page (homepage or a /contact-style subpage) and scan it
    for contact channels."""
    if not url:
        return _empty_scan_result()
    html = fetch_page(url)
    if html is None:
        return _empty_scan_result()
    return scan_page_for_contacts(html, own_username, f"website ({url})")


def contact_subpages(base_url: str) -> list[str]:
    """Typical contact-info pages worth an extra request on a plain website."""
    parts = urlsplit(base_url)
    if not parts.scheme or not parts.netloc:
        return []
    root = urlunsplit((parts.scheme, parts.netloc, "", "", ""))
    return [root + p for p in CONTACT_PAGE_PATHS]


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
MAILTO_RE = re.compile(r'mailto:([^"\'>\s]+)', re.IGNORECASE)
WA_ME_RE = re.compile(r"wa\.me/(\+?[\d\-\s()]{6,})", re.IGNORECASE)
WA_SEND_RE = re.compile(r"(?:api\.)?whatsapp\.com/send\?[^\s]*phone=(\+?[\d\-\s()]{6,})", re.IGNORECASE)
# Every absolute URL appearing anywhere in a fetched page (href attributes, plain text).
URL_IN_HTML_RE = re.compile(r'https?://[^\s"\'<>)]+')

TELEGRAM_DOMAINS = ("t.me", "telegram.me", "telegram.dog")
WHATSAPP_DOMAINS = ("wa.me", "whatsapp.com")
BOOKING_DOMAINS = ("calendly.com", "cal.com", "koalendar.com", "tidycal.com", "acuityscheduling.com")


def _domain_matches(url: str, domains: tuple[str, ...]) -> bool:
    """True if url's hostname is exactly one of `domains` or a subdomain of one.

    A plain substring check (`"cal.com" in url`) also matches unrelated hosts
    like "facebook.dlocal.com", or "t.me" matching "...content.medium.com":
    both seen in real deep-scan output. Comparing against the parsed hostname
    avoids that."""
    host = urlsplit(url if "://" in url else f"//{url}").hostname or ""
    host = host.lower()
    return any(host == d or host.endswith("." + d) for d in domains)


def extract_email(profile: Profile) -> str:
    """Prefer the provider's business/public email field; fall back to a bio regex."""
    if profile.provider_email:
        return profile.provider_email
    match = EMAIL_RE.search(profile.biography or "")
    return match.group(0) if match else ""


def extract_whatsapp(text: str) -> str:
    """Pull a phone number out of a wa.me / whatsapp.com link. WhatsApp has no usernames."""
    match = WA_ME_RE.search(text) or WA_SEND_RE.search(text)
    if not match:
        return ""
    digits = re.sub(r"\D", "", match.group(1))
    return f"+{digits}" if digits else ""


def classify_link(url: str) -> str:
    """Bucket a bio link as telegram / whatsapp / booking / aggregator / website."""
    if _domain_matches(url, TELEGRAM_DOMAINS):
        return "telegram"
    if _domain_matches(url, WHATSAPP_DOMAINS):
        return "whatsapp"
    if _domain_matches(url, BOOKING_DOMAINS):
        return "booking"
    # Aggregator entries are brand-name fragments, not full registrable domains
    # (an actor may hand back linktr.ee, linktree.com, or a country-TLD mirror),
    # so a substring check is intentional here: the only thing it gates is
    # "fetch this page and scan it", not a value reported as a contact.
    low = url.lower()
    if any(d in low for d in LINK_AGGREGATOR_DOMAINS):
        return "aggregator"
    return "website"


def extract_booking_and_website(urls: list[str]) -> tuple[str, str]:
    """Pick the first booking link and the first leftover link to report as the website."""
    booking = ""
    website = ""
    for url in urls:
        kind = classify_link(url)
        if kind == "booking" and not booking:
            booking = url
        elif kind == "website" and not website:
            website = url
    return booking, website


def _apply_scanned(profile: "Profile", scanned: dict, stats: dict, label: str) -> None:
    """Fill in whichever contact fields are still empty using a fetched page's results,
    crediting the funnel counter named by `label` (from_linktree / from_website)."""
    if not profile.email and scanned["email"]:
        profile.email = scanned["email"]
        stats[label] = stats.get(label, 0) + 1
    if not profile.whatsapp and scanned["whatsapp"]:
        profile.whatsapp = scanned["whatsapp"]
        stats[label] = stats.get(label, 0) + 1
    if not profile.telegram_username and scanned["telegram_username"]:
        profile.telegram_username = scanned["telegram_username"]
        profile.notes = f"telegram found via {label}"
        stats[label] = stats.get(label, 0) + 1
    if not profile.booking_link and scanned["booking_link"]:
        profile.booking_link = scanned["booking_link"]
        stats[label] = stats.get(label, 0) + 1
    for k, v in scanned["tg_stats"].items():
        stats[k] = stats.get(k, 0) + v


def _all_contacts_found(profile: "Profile") -> bool:
    return bool(profile.email and profile.whatsapp and profile.telegram_username
                and profile.booking_link)


def find_contacts(profile: Profile, deep_scan: bool = True) -> tuple[bool, dict]:
    """Try every contact channel (email, WhatsApp, Telegram, booking link, website).

    First pass is bio-only (free: no HTTP calls beyond t.me classification).
    If `deep_scan` is on and something is still missing, follow every link in
    externalUrls: link aggregators (Linktree/Taplink/...) and plain websites get
    fetched and scanned for contacts; plain websites also get up to
    MAX_EXTRA_WEBSITE_PAGES extra requests to typical contact pages.

    Populates the matching fields on `profile`. Returns (found_any, stats), where
    stats carries the Telegram bot/channel/personal breakdown plus a
    from_bio / from_linktree / from_website funnel breakdown for logging.
    """
    stats = {"candidates": 0, "bot": 0, "channel": 0, "unknown": 0, "personal": 0,
              "from_bio": 0, "from_linktree": 0, "from_website": 0}
    own = profile.username.lower()

    # pass 1: bio only
    profile.email = extract_email(profile)
    if profile.email:
        stats["from_bio"] += 1

    profile.whatsapp = extract_whatsapp(f"{profile.biography} {profile.external_url}")
    if profile.whatsapp:
        stats["from_bio"] += 1

    telegram, note, tg_stats = extract_telegram(profile)
    for key, value in tg_stats.items():
        stats[key] = stats.get(key, 0) + value
    if telegram:
        profile.telegram_username = telegram
        profile.notes = f"telegram {note}"
        stats["from_bio"] += 1

    profile.booking_link, profile.website = extract_booking_and_website(profile.external_urls)
    if profile.booking_link:
        stats["from_bio"] += 1

    # pass 2: follow every bio link and scan the page it points to
    if deep_scan and not _all_contacts_found(profile):
        extra_calls_left = MAX_EXTRA_WEBSITE_PAGES
        for url in profile.external_urls:
            if not url or _all_contacts_found(profile):
                break

            kind = classify_link(url)
            if kind in ("telegram", "whatsapp", "booking"):
                continue  # already extracted directly from the URL itself, no fetch needed

            if kind == "aggregator":
                scanned = resolve_link_aggregator(url, own)
                label = "from_linktree"
            else:
                scanned = resolve_website(url, own)
                label = "from_website"

            if not scanned["fetched"]:
                continue
            _apply_scanned(profile, scanned, stats, label)

            if kind == "website" and extra_calls_left > 0 and not _all_contacts_found(profile):
                for sub_url in contact_subpages(url):
                    if extra_calls_left <= 0 or _all_contacts_found(profile):
                        break
                    sub_scanned = resolve_website(sub_url, own)
                    extra_calls_left -= 1
                    if not sub_scanned["fetched"]:
                        continue
                    _apply_scanned(profile, sub_scanned, stats, "from_website")

    found_any = bool(
        profile.email or profile.whatsapp or profile.telegram_username
        or profile.booking_link or profile.website
    )
    return found_any, stats


def is_expert(profile: Profile) -> bool:
    text = profile.search_text
    if any(bad in text for bad in EXCLUDE_KEYWORDS):
        return False
    return any(kw in text for kw in EXPERT_KEYWORDS)


def passes_numeric(profile: Profile) -> bool:
    return (MIN_FOLLOWERS <= profile.followers <= MAX_FOLLOWERS
            and profile.following <= MAX_FOLLOWING)


def score_relevance(profile: Profile) -> int:
    """+1 per BONUS_KEYWORDS match in bio/name/category. Never rejects a profile."""
    text = profile.search_text
    return sum(1 for kw in BONUS_KEYWORDS if kw in text)


def _contact_summary(p: Profile) -> str:
    parts = []
    if p.email:
        parts.append("email")
    if p.whatsapp:
        parts.append("whatsapp")
    if p.telegram_username:
        parts.append(f"telegram=@{p.telegram_username}")
    if p.booking_link:
        parts.append("booking")
    if p.website:
        parts.append("website")
    return ",".join(parts) or "none"


class ApifySource:
    """Test-friendly source. Keyword discovery plus a detail-enrichment step."""

    SEARCH_ACTOR = "apify/instagram-scraper"          # $1.50 / 1k
    PROFILE_ACTOR = "apify/instagram-profile-scraper"  # $1.60 / 1k

    def __init__(self, token: str):
        from apify_client import ApifyClient
        self.client = ApifyClient(token)

    def _run(self, actor_id: str, run_input: dict) -> list[dict]:
        log.info("Apify: running %s", actor_id)
        try:
            run = self.client.actor(actor_id).call(run_input=run_input, logger=None)
        except TypeError:
            # older apify-client versions don't accept the logger kwarg
            run = self.client.actor(actor_id).call(run_input=run_input)
        if run is None:
            log.warning("Apify: run returned None for %s", actor_id)
            return []
        ds_id = run.default_dataset_id if hasattr(run, "default_dataset_id") else run["defaultDatasetId"]
        return list(self.client.dataset(ds_id).iterate_items())

    def discover(self, keyword: str, limit: int) -> list[dict]:
        # searchType="user" -> search for accounts by keyword. Some actor versions
        # already return bio/follower data; if not, collect() enriches via enrich().
        run_input = {
            "search": keyword,
            "searchType": "user",
            "searchLimit": limit,
            "resultsType": "details",
            "resultsLimit": limit,
            "addParentData": False,
        }
        return self._run(self.SEARCH_ACTOR, run_input)

    def enrich(self, usernames: list[str]) -> list[dict]:
        if not usernames:
            return []
        # The primary actor accepts a list of usernames; if a newer version expects
        # URLs instead, switch to {"directUrls": [f"https://instagram.com/{u}" for u in usernames]}.
        return self._run(self.PROFILE_ACTOR, {"usernames": usernames})


class HikerApiSource:
    """Volume source, backed by the official hikerapi SDK."""

    def __init__(self, token: str):
        from hikerapi import Client
        self.client = Client(token=token)

    def discover(self, keyword: str, limit: int) -> list[dict]:
        res = self.client.fbsearch_accounts_v3(query=keyword)
        users = res.get("users") if isinstance(res, dict) else res
        return (users or [])[:limit]

    def enrich(self, usernames: list[str]) -> list[dict]:
        out = []
        for u in usernames:
            try:
                out.append(self.client.user_by_username_v1(username=u))
            except Exception as e:
                log.debug("hiker enrich %s: %s", u, e)
            time.sleep(0.05)  # gentle throttling
        return out


def build_source(name: str):
    if name == "apify":
        token = os.getenv("APIFY_TOKEN")
        if not token:
            sys.exit("APIFY_TOKEN is not set.")
        return ApifySource(token)
    if name == "hiker":
        token = os.getenv("HIKER_TOKEN")
        if not token:
            sys.exit("HIKER_TOKEN is not set.")
        return HikerApiSource(token)
    sys.exit(f"Unknown source: {name}")


def collect(source, keywords: list[str], target: int,
            per_keyword: int = 30, exclude_markers: Optional[list[str]] = None,
            deep_scan: bool = True) -> list[Profile]:
    """Run the full pipeline over each keyword until `target` unique contacts are found."""
    exclude_markers = exclude_markers or []
    seen_ig: set[str] = set()      # dedup by Instagram username
    seen_tg: set[str] = set()      # dedup by Telegram handle
    results: list[Profile] = []

    for kw in keywords:
        if len(results) >= target:
            break
        log.info("Keyword: %r (collected %d/%d)", kw, len(results), target)

        try:
            raw_items = source.discover(kw, per_keyword)
        except Exception as e:
            log.warning("discover(%r) failed: %s", kw, e)
            continue

        log.info("  discover(%r) returned %d profiles", kw, len(raw_items))

        profiles = [p for p in (normalize_profile(d) for d in raw_items) if p]

        # If profiles came back without bio/follower data, fetch full details.
        need = [p.username for p in profiles if not p.biography or p.followers == 0]
        if need:
            try:
                enriched = {
                    ep.username.lower(): ep
                    for ep in (normalize_profile(d) for d in source.enrich(need))
                    if ep
                }
                profiles = [enriched.get(p.username.lower(), p) for p in profiles]
            except Exception as e:
                log.warning("enrich failed: %s", e)

        n_dup_ig = n_excluded = n_numeric_fail = n_expert_fail = 0
        n_no_contact = n_dup_tg = n_ok = 0
        n_tg_candidates = n_bot = n_channel = n_personal_found = 0
        n_email = n_whatsapp = n_telegram = n_booking = n_website = 0
        n_from_bio = n_from_linktree = n_from_website = 0

        for p in profiles:
            if len(results) >= target:
                break
            p.matched_keyword = kw
            if p.username.lower() in seen_ig:
                n_dup_ig += 1
                continue
            if matches_exclusion_markers(p, exclude_markers):
                n_excluded += 1
                continue
            if not passes_numeric(p):
                n_numeric_fail += 1
                continue
            if not is_expert(p):
                n_expert_fail += 1
                continue

            found, contact_stats = find_contacts(p, deep_scan=deep_scan)
            n_tg_candidates += contact_stats.get("candidates", 0)
            n_bot += contact_stats.get("bot", 0)
            n_channel += contact_stats.get("channel", 0)
            n_personal_found += contact_stats.get("personal", 0)
            n_from_bio += contact_stats.get("from_bio", 0)
            n_from_linktree += contact_stats.get("from_linktree", 0)
            n_from_website += contact_stats.get("from_website", 0)
            if not found:
                n_no_contact += 1
                continue

            if p.telegram_username:
                tg_key = p.telegram_username.lower()
                if tg_key in seen_tg:
                    n_dup_tg += 1
                    continue
                seen_tg.add(tg_key)

            p.relevance_score = score_relevance(p)
            seen_ig.add(p.username.lower())
            results.append(p)
            n_ok += 1
            n_email += bool(p.email)
            n_whatsapp += bool(p.whatsapp)
            n_telegram += bool(p.telegram_username)
            n_booking += bool(p.booking_link)
            n_website += bool(p.website)
            log.info("  + %-22s  %-30s  %d followers  score=%d",
                     p.username, _contact_summary(p), p.followers, p.relevance_score)

        log.info(
            "  funnel %r: profiles=%d dup_ig=%d excluded=%d numeric_fail=%d not_expert=%d | "
            "tg_candidates=%d bot/channel_dropped=%d personal_tg_found=%d dup_tg=%d "
            "no_contact=%d passed=%d | contacts: email=%d whatsapp=%d telegram=%d "
            "booking=%d website=%d | contact source: from_bio=%d from_linktree=%d "
            "from_website=%d",
            kw, len(profiles), n_dup_ig, n_excluded, n_numeric_fail, n_expert_fail,
            n_tg_candidates, n_bot + n_channel, n_personal_found, n_dup_tg,
            n_no_contact, n_ok, n_email, n_whatsapp, n_telegram, n_booking, n_website,
            n_from_bio, n_from_linktree, n_from_website,
        )

    results.sort(key=lambda p: p.relevance_score, reverse=True)
    return results


def _excel_safe(value: str) -> str:
    """Prefix with an apostrophe if Excel would otherwise read the value as a formula."""
    if value and value[0] in ("@", "=", "+", "-"):
        return "'" + value
    return value


def export_excel(profiles: list[Profile], path: str) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    rows = []
    for p in profiles:
        row = {}
        for attr, title in EXCEL_COLUMNS:
            value = getattr(p, attr)
            row[title] = _excel_safe(value) if isinstance(value, str) else value
        rows.append(row)
    df = pd.DataFrame(rows, columns=[title for _, title in EXCEL_COLUMNS])

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Leads")
        ws = xw.sheets["Leads"]

        widths = [20, 38, 22, 42, 11, 11, 20, 18, 14, 26, 16, 20, 30, 30, 30, 30]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        bio_col_idx = [attr for attr, _ in EXCEL_COLUMNS].index("biography") + 1
        wrap = Alignment(wrap_text=True, vertical="top")
        for cell in ws[get_column_letter(bio_col_idx)][1:]:
            cell.alignment = wrap

        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    log.info("Saved %d rows -> %s", len(df), path)


def main():
    ap = argparse.ArgumentParser(
        description="Instagram lead generation: extract contacts from expert profiles")
    ap.add_argument("--source", choices=["apify", "hiker"], default="apify")
    ap.add_argument("--limit", type=int, default=50, help="how many contacts to collect")
    ap.add_argument("--per-keyword", type=int, default=30, help="profiles to fetch per keyword")
    ap.add_argument("--out", default="instagram_leads.xlsx")
    ap.add_argument("--keywords", nargs="*", default=EXPERT_KEYWORDS,
                    help="override the default keyword list")
    ap.add_argument("--exclude-countries", nargs="*", default=[],
                    help="optional markers (city/country names, words) to exclude profiles by; "
                         "off by default")
    ap.add_argument("--deep-scan", dest="deep_scan", action="store_true", default=True,
                    help="follow bio links (link aggregators + websites) to look for contacts "
                         "when the bio itself has none (default: on)")
    ap.add_argument("--no-deep-scan", dest="deep_scan", action="store_false",
                    help="only look for contacts in the bio itself, skip fetching any links")
    args = ap.parse_args()

    source = build_source(args.source)
    log.info("Source: %s | target: %d contacts | deep_scan: %s",
              args.source, args.limit, args.deep_scan)

    profiles = collect(source, args.keywords, args.limit, args.per_keyword,
                        args.exclude_countries, deep_scan=args.deep_scan)

    if not profiles:
        log.warning("Nothing collected, check token/filters/actor limits.")
        return

    export_excel(profiles, args.out)

    log.info(
        "Overall contacts: email=%d whatsapp=%d telegram=%d booking=%d website=%d",
        sum(bool(p.email) for p in profiles),
        sum(bool(p.whatsapp) for p in profiles),
        sum(bool(p.telegram_username) for p in profiles),
        sum(bool(p.booking_link) for p in profiles),
        sum(bool(p.website) for p in profiles),
    )
    log.info("Done. Total unique contacts: %d", len(profiles))
    for p in profiles[:5]:
        log.info("  %-20s %-30s matched=%s", p.username, _contact_summary(p), p.matched_keyword)


if __name__ == "__main__":
    main()
